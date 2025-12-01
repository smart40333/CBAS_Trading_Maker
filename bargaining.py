import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from db_access import get_customer_info, get_customer_inventory
from format_utils import format_date, convert_to_chinese_amount, strip_trailing_zeros, next_business_day, strip_whitespace
from file_reader import read_quote_excel, read_vip_list, read_vip_quote
import pyodbc
from decimal import Decimal, ROUND_HALF_UP



def process_bargain_records(bargain_df: pd.DataFrame, df_quote: pd.DataFrame) -> pd.DataFrame:
    """處理議價交易資料，回傳處理後的 DataFrame（不操作 UI）。"""
    try:
        if bargain_df is None or bargain_df.empty:
                return pd.DataFrame()
                
        # 準備 CB 基本資料
        cb_info = df_quote[['CB代號', 'CB名稱']] if ('CB代號' in df_quote.columns and 'CB名稱' in df_quote.columns) else pd.DataFrame(columns=['CB代號', 'CB名稱'])

        # 取得需要查詢的客戶
        db_cusid_length = 12
        cusids = bargain_df['客戶ID'].astype(str).unique().tolist() if '客戶ID' in bargain_df.columns else []
        cusids_padded = [cusid.ljust(db_cusid_length) for cusid in cusids]
        df_cus_info = get_customer_info(cusids_padded)

        # 合併
        df_bargain_con = bargain_df.copy()
        df_bargain_con = df_bargain_con.merge(df_cus_info, left_on='客戶ID', right_on='CUSID', how='left')
        if 'CB名稱' in df_bargain_con.columns:
            df_bargain_con = df_bargain_con.drop('CB名稱', axis=1)
        df_bargain_con = df_bargain_con.merge(cb_info, on='CB代號', how='left') if not cb_info.empty else df_bargain_con
                
        # 補充欄位
        df_bargain_con.loc[:, '客戶名稱'] = df_bargain_con.get('CUSNAME')
        df_bargain_con.loc[:, '銀行'] = df_bargain_con.get('BNKNAME')
        df_bargain_con.loc[:, '分行'] = df_bargain_con.get('BNKBRH')
        df_bargain_con.loc[:, '銀行帳號'] = df_bargain_con.get('BNKACTNO')
        df_bargain_con.loc[:, '集保帳號'] = df_bargain_con.get('CENTERNO')
        df_bargain_con.loc[:, '通訊地址'] = df_bargain_con.get('ADDRESS2')
        #向量化操作（更高效）
        try:
            df_bargain_con.loc[:, '交割日期'] = df_bargain_con.apply(
                lambda row: next_business_day(
                    pd.to_datetime(row['成交日期'], format='%Y%m%d'), 
                    int(row['T+?交割'])
                ).strftime('%Y%m%d'), 
                axis=1
            )
        except Exception as e:
            print(f"計算交割日期發生錯誤: {e}")
            df_bargain_con.loc[:, '交割日期'] = ''
            
        # 計算金額
        df_bargain_con['議價張數'] = pd.to_numeric(df_bargain_con.get('議價張數'), errors='coerce')
        df_bargain_con['議價價格'] = pd.to_numeric(df_bargain_con.get('議價價格'), errors='coerce').astype(float)
        df_bargain_con['議價金額'] = (df_bargain_con['議價張數'] * df_bargain_con['議價價格'] * 1000).round().astype('Int64')
        df_bargain_con['議價金額'] = df_bargain_con['議價金額'].apply(lambda x: f"{int(x):,}" if pd.notna(x) else '')
        df_bargain_con['備註'] = '議價交易日期' + df_bargain_con['成交日期'].astype(str) + ' ' + df_bargain_con['錄音時間'].astype(str) + '議價' + df_bargain_con['買/賣'] + df_bargain_con['CB代號'].astype(str) + df_bargain_con['CB名稱'] + df_bargain_con['議價張數'].astype(str) + '張，按' + df_bargain_con['參考價'] + df_bargain_con['議價價格'].astype(str) + '進行，與市場價格未有差異'

        # 欄位順序
        final_columns = [
            '單據編號', '成交日期', 'T+?交割', '買/賣','客戶ID', 'CB代號', '議價張數', '議價價格', '參考價', '錄音時間', '交割日期',
            '客戶名稱', 'CB名稱', '議價金額', '備註', '銀行', '分行', '銀行帳號', '集保帳號', '通訊地址'
        ]
        for col in final_columns:
            if col not in df_bargain_con.columns:
                df_bargain_con[col] = ''
    except Exception as e:
        import traceback
        import sys
        exc_type, exc_value, exc_tb = sys.exc_info()
        tb = traceback.extract_tb(exc_tb)
        if tb:
            last_tb = tb[-1]
            print(f"處理議價交易資料時發生錯誤 (行數 {last_tb.lineno}):\n{e}\n詳細內容:\n{traceback.format_exc()}")
        else:
            print(f"處理議價交易資料時發生錯誤: {e}")
        return pd.DataFrame()

    return df_bargain_con[final_columns].copy()


def preserve_special_symbols(ws, buy_or_sell) -> None:
    #for cell_addr in ['A4', 'A5', 'G4']:
    #    cell = ws[cell_addr]
    #    if cell.value and isinstance(cell.value, str) and cell.value.startswith('R'):
    #        cell.value = cell.value.replace('R', '☑ ', 1)
    if buy_or_sell == '買':
        ws['A4'] = '☑ 自營買進'
        ws['A5'] = '☐ 自營賣出'
        ws['G4'] = '☑ 營業處所議價'
    elif buy_or_sell == '賣':
        ws['A4'] = '☐ 自營買進'
        ws['A5'] = '☑ 自營賣出'
        ws['G4'] = '☑ 營業處所議價'


def save_as_pdf(excel_path, pdf_path=None):
    """將議價交易 Excel 檔案轉換為 PDF"""
    try:
        import win32com.client
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        workbook = excel.Workbooks.Open(excel_path)
        
        if pdf_path is None:
            pdf_path = excel_path.replace('.xlsx', '.pdf')
        
        workbook.ExportAsFixedFormat(0, pdf_path)
        workbook.Close(False)
        excel.Quit()
        
        print(f"✓ 議價交易 PDF 已產生：{pdf_path}")
        return pdf_path
        
    except Exception as e:
        print(f"❌ 轉換 PDF 失敗：{e}")
        return None


def generate_settlement_voucher(row: dict) -> str:
    """生成櫃檯買賣合併債券給付結算憑單暨交付清單，回傳輸出檔路徑。"""
    template_path = r"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\議價模板_給付.xlsx"
    wb = load_workbook(template_path)
    ws = wb['給付']
    ws['B4'] = row.get('客戶名稱', '')
    ws['B5'] = row.get('集保帳號', '')
    ws['I4'] = format_date(row.get('成交日期', ''))
    ws['I5'] = format_date(row.get('交割日期', ''))
    ws['A7'] = row.get('單據編號', '')
    ws['C7'] = f"{row.get('CB代號', '')}{row.get('CB名稱', '')}"
    ws['E7'] = row.get('議價張數', '')
    ws['I7'] = row.get('議價金額', '')

    if row.get('買/賣', '') == '買':
        ws['B7'] = '買斷'
        ws['G7'] = '收'
        ws['F15'] = f"{row.get('銀行', '')}{row.get('分行', '')}"
        ws['F17'] = row.get('銀行帳號', '')
        ws['B15'] = ''
        ws['B17'] = ''
        ws['I15'] = row.get('議價金額', '')
        ws['I18'] = row.get('議價金額', '')
        ws['I19'] = row.get('議價金額', '')
    elif row.get('買/賣', '') == '賣':
        ws['B7'] = '賣斷'
        ws['G7'] = '付'
        ws['F15'] = ''
        ws['F17'] = ''
        ws['B15'] = f"{row.get('銀行', '')}{row.get('分行', '')}"
        ws['B17'] = row.get('銀行帳號', '')
        ws['D15'] = row.get('議價金額', '')
        ws['D18'] = row.get('議價金額', '')
        ws['D19'] = row.get('議價金額', '')
    
    try:
        ws['D7'] = int(float(row.get('議價張數', 0))) * 100000
    except Exception:
        ws['D7'] = 0
    
    tday = datetime.now().strftime('%Y%m%d')
    output_dir = rf"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\{tday}"
    os.makedirs(output_dir, exist_ok=True)
    filename = f"給付憑證_{row.get('客戶名稱', 'Unknown')}_{row.get('單據編號', '')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    save_as_pdf(filepath)
    return filepath


def generate_trading_slip(row: dict) -> str:
    """生成櫃台買賣可轉債買賣成交單，回傳輸出檔路徑。"""
    template_path = r"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\議價模板_買賣.xlsx"
    wb = load_workbook(template_path)
    ws = wb['買賣']
    buy_or_sell = row.get('買/賣', '')
    preserve_special_symbols(ws, buy_or_sell)
    ws['A7'] = f"交易對手：{row.get('客戶名稱', '')}"
    ws['D7'] = f"集保帳號：{row.get('集保帳號', '')}"
    ws['I4'] = row.get('單據編號', '')
    ws['I6'] = format_date(row.get('成交日期', ''))
    ws['I7'] = format_date(row.get('交割日期', ''))
    ws['B8'] = row.get('通訊地址', '')
    ws['I8'] = row.get('客戶ID', '')
    ws['A12'] = f"{row.get('CB代號', '')}{row.get('CB名稱', '')}"
    ws['B12'] = row.get('議價張數', '')
    ws['C12'] = row.get('議價價格', '')
    ws['E12'] = row.get('議價金額', '')
    ws['I12'] = row.get('議價金額', '')
    ws['I17'] = row.get('議價金額', '')
    amount = str(row.get('議價金額', '')).replace(',', '').replace(' ', '')
    ws['B18'] = convert_to_chinese_amount(amount)
    tday = datetime.now().strftime('%Y%m%d')
    output_dir = rf"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\{tday}"
    os.makedirs(output_dir, exist_ok=True)
    filename = f"成交單_{row.get('客戶名稱', 'Unknown')}_{row.get('單據編號', '')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    save_as_pdf(filepath)
    return filepath

def generate_bargain_upload_file(df_bargain):
    """產生議價交易上傳檔"""
    df_template = pd.read_excel(r"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\議價檔ASBARG上傳檔.xlsx")
    df_fill = df_template.iloc[0:0]
    df_fill['TXDATE'] = df_bargain['成交日期']
    df_fill['SETDAT'] = df_bargain['交割日期']
    df_fill['ORDERNO'] = df_bargain['單據編號']
    df_fill['CUSID'] = df_bargain['客戶ID']
    df_fill['BRKID'] = df_bargain['集保帳號'].str[:4]
    df_fill['ACCTNO'] = df_bargain['集保帳號'].str[4:]
    df_fill['TXBS'] = np.where(df_bargain['買/賣'] == '買', 'B', 'S')
    df_fill['STOCK'] = df_bargain['CB代號']
    df_fill['MTHQTY'] = df_bargain['議價張數']
    df_fill['PRICE'] = df_bargain['議價價格']
    df_fill['MTHAMT'] = df_bargain['議價金額'].str.replace(',', '').replace(' ', '')
    df_fill['OFFSET'] = 'Y'
    df_fill['RECUSER'] = '10112'
    df_fill['RECTIME'] = df_bargain['錄音時間']
    df_fill.to_excel(r"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\議價檔ASBARG上傳檔.xlsx", index=False)
    return df_fill
    
def calculate_new_trade_batch(trade_data: pd.DataFrame) -> pd.DataFrame:
    """統一的新作買進批次計算 - 使用獨立模組讀取資料"""
    try:
        # 讀取所需資料
        df_vip_quote_all = read_vip_quote()
        df_vip_list = read_vip_list()
        df_cus_inventory = get_customer_inventory()
        df_quote = read_quote_excel()

        df_trade = trade_data.copy()
        
        if df_cus_inventory.empty or 'STORQTY' not in df_cus_inventory.columns:
            df_cus_inventory = pd.DataFrame(columns=['CUSID', 'STORQTY'])
        
        # 合併報價資訊
        df_trade = pd.merge(df_trade, df_quote[['CB代號', 'CB名稱', '賣回日', '賣回價', '履約利率', '低履約利率', '選擇權到期日', '波動度']], on='CB代號', how='left')
        if 'CB名稱_x' in df_trade.columns:
                df_trade['CB名稱'] = df_trade['CB名稱_x']
        
    # 合併客戶庫存
        if not df_cus_inventory.empty:
            df_trade = pd.merge(df_trade, df_cus_inventory[['CUSID', 'STORQTY']], left_on='客戶ID', right_on='CUSID', how='left')
        else:
            df_trade['STORQTY'] = 0
            
        if 'STORQTY' not in df_trade.columns:
            df_trade['STORQTY'] = 0
        
        if '今履約利率' in df_trade.columns:
            df_trade['履約利率'] = df_trade['今履約利率']
            df_trade['低履約利率'] = df_trade['今履約利率']
        
    # 數值轉換
        df_trade['STORQTY'] = pd.to_numeric(df_trade['STORQTY'], errors='coerce').fillna(0)
        df_trade['成交均價'] = pd.to_numeric(df_trade['成交均價'], errors='coerce')
        df_trade['成交張數'] = pd.to_numeric(df_trade['成交張數'], errors='coerce')
        df_trade['履約利率'] = pd.to_numeric(df_trade['履約利率'], errors='coerce')
        df_trade['低履約利率'] = pd.to_numeric(df_trade['低履約利率'], errors='coerce')
        df_trade['賣回價'] = pd.to_numeric(df_trade['賣回價'], errors='coerce')
    
    # 計算年期
        settle = next_business_day(pd.Timestamp.today().normalize(), 2).normalize()
        sellback = pd.to_datetime(df_trade['賣回日'], format='%Y%m%d').dt.normalize()
        df_trade['年期_app'] = ((sellback - settle).dt.days + 1) / 365
        
    # 特殊ID處理
        special_ids = ['H122699830', 'H123326603', 'P220839691']
        df_trade['最終利率'] = np.nan
        df_trade['最終手續費'] = np.nan
        
    # 1. 檢查特殊報價
        if not df_vip_quote_all.empty and all(col in df_vip_quote_all.columns for col in ['客戶ID', 'CB代號', '利率%', '手續費']):
            vip_quote_key = df_vip_quote_all['客戶ID'].astype(str) + '_' + df_vip_quote_all['CB代號'].astype(str)
            vip_quote_dict = df_vip_quote_all.set_index(vip_quote_key)[['利率%', '手續費']].to_dict('index')
            
            for idx, row in df_trade.iterrows():
                trade_key = str(row['客戶ID']) + '_' + str(row['CB代號'])
                if trade_key in vip_quote_dict:
                    vip_info = vip_quote_dict[trade_key]
                    if pd.notna(vip_info.get('利率%')):
                        df_trade.loc[idx, '最終利率'] = vip_info['利率%']
                    if pd.notna(vip_info.get('手續費')):
                        df_trade.loc[idx, '最終手續費'] = vip_info['手續費']
        
    # 2. 檢查VIP名單
        if not df_vip_list.empty and all(col in df_vip_list.columns for col in ['客戶ID', '不限張數低手續費', '不限張數低利率']):
            vip_list_dict = df_vip_list.set_index('客戶ID')[['不限張數低手續費', '不限張數低利率']].to_dict('index')
            
            for idx, row in df_trade.iterrows():
                cus_id = str(row['客戶ID'])
                if pd.isna(df_trade.loc[idx, '最終利率']) or pd.isna(df_trade.loc[idx, '最終手續費']):
                    if cus_id in vip_list_dict:
                        vip_info = vip_list_dict[cus_id]
                        if pd.isna(df_trade.loc[idx, '最終利率']) and vip_info.get('不限張數低利率') == 'Y':
                            df_trade.loc[idx, '最終利率'] = row['低履約利率']
                        if pd.isna(df_trade.loc[idx, '最終手續費']) and vip_info.get('不限張數低手續費') == 'Y':
                            df_trade.loc[idx, '最終手續費'] = 100
        
    # 3. 檢查特殊ID
        for idx, row in df_trade.iterrows():
            cus_id = str(row['客戶ID'])
            need_process = (pd.isna(df_trade.loc[idx, '最終利率']) or pd.isna(df_trade.loc[idx, '最終手續費']) or str(df_trade.loc[idx, '最終手續費']) == 'nan' or str(df_trade.loc[idx, '最終手續費']) == '')
            if need_process and cus_id in special_ids:
                if pd.isna(df_trade.loc[idx, '最終利率']):
                    df_trade.loc[idx, '最終利率'] = row['低履約利率']
                if (pd.isna(df_trade.loc[idx, '最終手續費']) or str(df_trade.loc[idx, '最終手續費']) == 'nan' or str(df_trade.loc[idx, '最終手續費']) == ''):
                            df_trade.loc[idx, '最終手續費'] = 60
        
    # 4. 基礎規則
        for idx, row in df_trade.iterrows():
            if pd.isna(df_trade.loc[idx, '最終利率']):
                df_trade.loc[idx, '最終利率'] = row['低履約利率'] if row['STORQTY'] >= 200 else row['履約利率']
            if pd.isna(df_trade.loc[idx, '最終手續費']):
                if row['STORQTY'] >= 200 or row['成交張數'] >= 10:
                        df_trade.loc[idx, '最終手續費'] = 100
                elif row.get('SRC', '') == 'E':
                    df_trade.loc[idx, '最終手續費'] = 110
                else:
                    df_trade.loc[idx, '最終手續費'] = 150
        
    # 最終處理
        df_trade['最終手續費'] = pd.to_numeric(df_trade['最終手續費'], errors='coerce').astype(int)
        df_trade['最終利率'] = pd.to_numeric(df_trade['最終利率'], errors='coerce')
        df_trade['成交均價'] = np.round(df_trade['成交均價'], 11)
        
        def calc_unit_premium(df_trade):
            prem100 = (
                df_trade['最終利率'] * df_trade['年期_app']
                - (df_trade['賣回價'] - 100)
                + df_trade['最終手續費'] / 1000
            )

            # 權利金百元價：固定到 2 位
            df_trade['權利金百元價'] = prem100.apply(
                lambda x: float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            )

            # 單位權利金 = (成交均價 - 100) + 權利金百元價
            # 成交均價不動，權利金是精確的兩位
            df_trade['單位權利金'] = (
                (df_trade['成交均價'] - 100) + df_trade['權利金百元價']
            )
            return df_trade


        df_trade = calc_unit_premium(df_trade)
        
        df_trade['權利金總額'] = np.round(df_trade['單位權利金'] * df_trade['成交張數'] * 1000, 0).astype(int)
        df_trade['履約利率%'] = df_trade['最終利率']
        #問題出在strip_trailing_zeros
        #df_trade['單位權利金'] = df_trade['單位權利金'].astype(str)
        #print(df_trade['單位權利金'])
        df_trade['單位權利金'] = df_trade['單位權利金'].apply(lambda x: strip_trailing_zeros(x))
        print(df_trade['單位權利金'])
        df_trade['成交均價'] = df_trade['成交均價'].apply(lambda x: strip_trailing_zeros(x))
        return df_trade
        
    except Exception as e:
        print(f"批量計算發生錯誤: {e}")
        return trade_data
 
def bargain_sell(bargain_data_sell):
    """議價賣出功能"""
    try:
        all_sell_data = []
        
        for index, row in bargain_data_sell.iterrows():
            cus_id = row['客戶ID']
            cb_code = row['CB代號']
            exercise_qty = row['議價張數']
            settlement_date = row['交割日期']
            trade_date = row['成交日期']
            price = row['議價價格']
            
            # 獲取履約契約資料
            df_exe = fetch_exercise_contracts(cus_id, cb_code, trade_date, exercise_qty, price, settlement_date)
            
            if not df_exe.empty:
                all_sell_data.append(df_exe)
        
        # 合併所有賣出資料
        if all_sell_data:
            df_all_sell = pd.concat(all_sell_data, ignore_index=True)
            return df_all_sell
        else:
            return pd.DataFrame()
            
    except Exception as e:
        print(f"議價賣出處理時發生錯誤: {e}")
        return pd.DataFrame()
        
def fetch_exercise_contracts(cus_id: str, cb_code: str, trade_date: str, exercise_qty: str, price: str, settlement_date: str) -> pd.DataFrame:
    """從資料庫獲取履約契約資訊並進行智能排序分配"""
    exercise_qty = int(exercise_qty)
    conn = pyodbc.connect('DSN=PSCDB', UID='FSP631', PWD='FSP631')
    
    # 補齊客戶ID到12位
    db_cusid_length = 12
    cus_id_padded = cus_id.ljust(db_cusid_length)
    
    # 查詢該客戶的現有契約
    sql_query = f"""
        SELECT 
            CUSID,
            PRDID,
            CBCODE,
            STORQTY,
            TRDATE,
            PERRATE,
            CBTPPRI,
            CBTPDT
        FROM FSPFLIB.ASPROD 
        WHERE CUSID = '{cus_id_padded}' 
        AND CBCODE = '{cb_code}' 
        AND STORQTY > 0
    """
    
    df_contracts = strip_whitespace(pd.read_sql(sql_query, conn))
    df_contracts.columns = ['客戶ID', '原單契約編號', 'CB代號', '庫存張數', '原成交日期', '原利率', '賣回價', '賣回日']
    
    if df_contracts.empty:
        conn.close()
        print(f"沒有找到符合條件的契約")
        return pd.DataFrame()
    
    # 獲取客戶名稱
    df_cusname = strip_whitespace(pd.read_sql(f"SELECT CUSID, CUSNAME FROM FSPFLIB.FSPCS0M WHERE CUSID = '{cus_id_padded}'", conn))
    df_quote = read_quote_excel()
    # 獲取CB名稱和相關資訊
    cb_info = df_quote[df_quote['CB代號'] == cb_code]

    conn.close()
    
    # 合併資料
    if not df_cusname.empty:
        df_contracts = pd.merge(df_contracts, df_cusname[['CUSID', 'CUSNAME']], left_on='客戶ID', right_on='CUSID', how='left')
        df_contracts['客戶名稱'] = df_contracts['CUSNAME']
    else:
        df_contracts['客戶名稱'] = ''
    
    if not cb_info.empty:
        df_contracts['CB名稱'] = cb_info.iloc[0]['CB名稱']
    else:
        df_contracts['CB名稱'] = ''
    
    # 數據類型轉換
    df_contracts['原利率'] = pd.to_numeric(df_contracts['原利率'], errors='coerce').fillna(0)
    df_contracts['庫存張數'] = pd.to_numeric(df_contracts['庫存張數'], errors='coerce').fillna(0).astype(int)
    df_contracts['賣回價'] = pd.to_numeric(df_contracts['賣回價'], errors='coerce').fillna(100)
    
    # 轉換日期格式
    df_contracts['原成交日期'] = pd.to_datetime(df_contracts['原成交日期'], format='%Y%m%d', errors='coerce')
    df_contracts['賣回日'] = pd.to_datetime(df_contracts['賣回日'], format='%Y%m%d', errors='coerce')
    
    # 智能排序：原利率(高→低)、成交日期(早→晚)、庫存張數(少→多)
    df_contracts_sorted = df_contracts.sort_values(
        by=['原利率', '原成交日期', '庫存張數'], 
        ascending=[False, True, True]
    ).reset_index(drop=True)
    
    # 創建 df_exe 進行履約分配
    df_exe = create_exercise_allocation(df_contracts_sorted, exercise_qty, settlement_date)
    
    # 檢查 df_exe 是否為空
    if df_exe.empty:
        print("沒有可履約的契約")
        return pd.DataFrame()
    
    # 檢查是否有庫存不足的情況
    total_exercised = df_exe['此契約履約張數'].sum() if not df_exe.empty else 0
    shortage = exercise_qty - total_exercised
    if shortage > 0:
        print(f"注意: 客戶可履約張數不足，缺少 {shortage} 張")
    
    # 轉換為顯示格式
    df_exe['履約張數'] = df_exe['此契約履約張數']
    df_exe['成交均價'] = price
    df_exe['賣出金額'] = (df_exe['成交均價'].astype(float) * df_exe['履約張數'].astype(int) * 1000).astype(int)
    df_exe['履約後剩餘張數'] = df_exe['原庫存張數'].astype(int) - df_exe['此契約履約張數'].astype(int)

    sell_columns = [
    '原單契約編號', '客戶ID', '客戶名稱', 'CB名稱', 'CB代號', '交易日期', '交割日期',
    '履約張數', '成交均價', '履約利率', '提前履約賠償金', '履約價', 
    '選擇權交割單價', '交割總金額', '錄音時間', '解約類別', '履約方式'
        ]

    df_exe['履約價'] = df_exe['履約價'].astype(float)
    df_exe['履約利率'] = df_exe['原利率']
    df_exe['履約張數'] = df_exe['履約張數'].astype(int)

    df_exe['選擇權交割單價'] = df_exe['成交均價'].astype(float) - df_exe['履約價'].astype(float)
    df_exe['交割總金額'] = df_exe['履約張數'].astype(int) * df_exe['選擇權交割單價'].astype(float) * 1000
    df_exe['提前履約賠償金'] = '0'
    df_exe['交易日期'] = trade_date
    df_exe['交割日期'] = settlement_date
    
    # 添加缺失的欄位
    df_exe['錄音時間'] = ''

    # 確保所有必要的欄位都存在
    for col in sell_columns:
        if col not in df_exe.columns:
            df_exe[col] = ''

    # 重新排列欄位順序
    df_exe = df_exe[sell_columns]
    
    print(f"已將 {len(df_exe)} 筆實物履約添加到提解賣出分頁！")
    return df_exe

def create_exercise_allocation(df_contracts_sorted, exercise_qty, settlement_date):
    """創建 df_exe 進行履約分配"""
    df_exe = df_contracts_sorted.copy()
    # Convert settlement_date from string to timestamp if it's a string
    if isinstance(settlement_date, str):
        settlement_date = pd.to_datetime(settlement_date, format='%Y%m%d')

    
    try:
        tday = datetime.now().strftime('%Y%m%d')
        df_had_sold = pd.read_csv(rf"\\10.72.228.83\CBAS_Logs\{tday}\AfterClose\ASCCSV02.csv")
        df_exe = pd.merge(df_exe, df_had_sold, left_on='原單契約編號', right_on='PRDID', how='left')
        df_exe['今日賣出張數'] = df_exe['DEUQTY'].fillna(0)  # 填充 NaN 值
    except Exception as e:
        print(f"讀取已賣出資料時發生錯誤: {e}")
        df_exe['今日賣出張數'] = 0

    try:
        # 添加履約相關欄位
        df_exe['此契約履約張數'] = 0
        df_exe['原庫存張數'] = df_exe['庫存張數'].copy()
        df_exe['剩餘張數'] = df_exe['庫存張數'].copy() - df_exe['今日賣出張數']
        df_exe['履約價'] = 0.0
        df_exe['年期'] = 0.0
        
        remaining_qty = exercise_qty
        
        for idx, row in df_exe.iterrows():
            if remaining_qty <= 0:
                break
            
            available_qty = int(row['庫存張數'])
            exercise_this_contract = min(remaining_qty, available_qty)
            
            # 更新履約張數
            df_exe.at[idx, '此契約履約張數'] = exercise_this_contract
            remaining_contract_qty = available_qty - exercise_this_contract
            df_exe.at[idx, '剩餘張數'] = remaining_contract_qty
            df_exe.at[idx, '解約類別'] = '2' if remaining_contract_qty == 0 else '1'
            df_exe.at[idx, '履約方式'] = '2'
            
            # 計算年期 = (賣回日 - 交割日) / 365
            if pd.notna(row['賣回日']):
                years = ((row['賣回日'] - settlement_date).days + 1)/ 365
                df_exe.at[idx, '年期'] = max(years, 0)  # 確保年期不為負
            else:
                df_exe.at[idx, '年期'] = 0
            
            # 計算履約價 = round(賣回價 - (100 * 年期 * 原利率), 2)
            exercise_price = round(
                row['賣回價'] - (100 * df_exe.at[idx, '年期'] * row['原利率']/100), 2
            )
            df_exe.at[idx, '履約價'] = exercise_price
            
            remaining_qty -= exercise_this_contract
            
            print(f"契約{idx+1}: 原利率={row['原利率']:.4f}, 可用={available_qty}張, 履約={exercise_this_contract}張, 履約價={exercise_price:.2f}")
        
        if remaining_qty > 0:
            print(f"⚠️  客戶可履約張數不足，缺少 {remaining_qty} 張")
        else:
            print(f"✅ 履約分配完成！")
        
        # 只返回有履約的契約
        df_exe_filtered = df_exe[df_exe['此契約履約張數'] > 0].reset_index(drop=True)
    except Exception as e:
        print(f"履約分配時發生錯誤: {e}")
        # 當發生錯誤時，返回空的 DataFrame 而不是不完整的 df_exe
        return pd.DataFrame()

    return df_exe_filtered



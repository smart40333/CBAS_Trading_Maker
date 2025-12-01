import os
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from db_access import get_clearing_detail
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
import win32com.client
from envs import trade_notice_dir


KAI_FONT_NAME = "DFKai-SB"  # 標楷體
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="E6F2FF")


def _set_col_widths(ws) -> None:
    column_widths = {
        1: 18,   # A
        2: 12,   # B
        3: 12,   # C
        4: 14,   # D
        5: 16,   # E
        6: 14,   # F
        7: 12,   # G
        8: 18,   # H
        9: 16,   # I
        10: 18,  # J
        11: 18,  # K
        12: 18,  # L (加宽以容纳"交割總金額"等内容)
        13: 14,  # M
        14: 18,  # N
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _thin_border() -> Border:
    thin = Side(border_style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _section_title(ws, row_idx: int, text: str) -> int:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=14)
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.font = Font(name=KAI_FONT_NAME, bold=True, size=14)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = WHITE_FILL
    return row_idx + 1


def _write_table(ws, start_row: int, headers: List[str], rows: List[List[Any]], header_fill: PatternFill = WHITE_FILL) -> int:
    border = _thin_border()

    # Header
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col, value=header)
        c.font = Font(name=KAI_FONT_NAME, bold=True, size=13)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        c.fill = header_fill

    # Rows
    for r_offset, data_row in enumerate(rows, start=1):
        for c_offset, value in enumerate(data_row, start=1):
            v = None if (isinstance(value, str) and value == "") else value
            if isinstance(v, float) and (np.isnan(v)):
                v = None
            c = ws.cell(row=start_row + r_offset, column=c_offset, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            c.font = Font(name=KAI_FONT_NAME, size=13)
            c.fill = WHITE_FILL

    return start_row + 1 + len(rows) + 1


def _write_merged_cell(ws, row: int, col_start: int, col_end: int, value: str, font_size: int = 11, bold: bool = False, alignment: str = "left") -> None:
    """輔助函數：寫入合併儲存格"""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = Font(name=KAI_FONT_NAME, size=font_size, bold=bold)
    cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=True)
    cell.fill = WHITE_FILL


def _convert_excel_to_pdf(excel_path: str, pdf_output_path: str) -> Optional[str]:
    """
    將Excel文件轉換為PDF
    
    Args:
        excel_path: Excel文件的完整絕對路徑
        pdf_output_path: PDF輸出目錄的完整絕對路徑
        
    Returns:
        PDF文件路徑，如果轉換失敗則返回None
    """
    try:
        # 從 excel_path 提取文件名，並改為 .pdf 副檔名
        excel_filename = os.path.basename(excel_path)  # 例如: "filename.xlsx"
        pdf_filename = os.path.splitext(excel_filename)[0] + '.pdf'  # 例如: "filename.pdf"
        pdf_path = os.path.join(pdf_output_path, pdf_filename)
        
        # 使用絕對路徑
        excel_path = os.path.abspath(excel_path)
        pdf_path = os.path.abspath(pdf_path)
        
        # 創建Excel應用程序對象
        excel = win32com.client.Dispatch("Excel.Application")
        # 某些環境（例如受限權限/服務帳戶）不允許設置 Visible/DisplayAlerts
        # 改為嘗試設置為整數（0/1），失敗則忽略
        try:
            excel.Visible = 0
        except Exception:
            pass
        try:
            excel.DisplayAlerts = 0
        except Exception:
            pass
        
        try:
            # 打開工作簿
            wb = excel.Workbooks.Open(excel_path)
            
            # 保存為PDF (0 = xlTypePDF)
            wb.ExportAsFixedFormat(0, pdf_path)
            
            # 關閉工作簿
            wb.Close(SaveChanges=False)
            
            return pdf_path
            
        finally:
            # 確保Excel應用程序被關閉
            excel.Quit()
            
    except Exception as e:
        print(f"Excel轉PDF時發生錯誤: {e}")
        return None


def _mask_account_number(account_num: str) -> str:
    """
    遮蔽账号中间部分，只保留前3码和后4码
    例如: 12345678901234 -> 123*******1234
    """
    if not account_num or not isinstance(account_num, str):
        return account_num
    
    account_num = str(account_num).strip()
    if len(account_num) <= 7:  # 如果长度<=7，不遮蔽
        return account_num
    
    # 保留前3码和后4码，中间用*替代
    masked = account_num[:3] + '*' * (len(account_num) - 7) + account_num[-4:]
    return masked


def _format_number_with_comma(value) -> str:
    """
    将数字格式化为千分位字符串
    例如: 1234567 -> 1,234,567
    """
    if pd.isna(value) or value is None:
        return None
    
    try:
        # 尝试转换为数字
        num = float(value)
        if np.isnan(num):
            return None
        # 转换为整数并格式化
        return f"{int(round(num)):,}"
    except (ValueError, TypeError):
        return value


def _add_logo(ws, logo_path: str, row: int = 1) -> None:
    """輔助函數：添加Logo並置中"""
    # 檢查圖片文件是否存在
    if not os.path.exists(logo_path):
        print(f"警告：Logo圖片路徑不存在: {logo_path}")
        return
    
    try:
        # 載入圖片
        img = XLImage(logo_path)
        scale = 0.9
        img.width = int(598 * scale)
        img.height = int(61 * scale)
        
        # 優先使用錨點方式實現置中
        try:
            # 計算合併儲存格A1:N1的總寬度（像素）
            # Excel列寬單位轉換：1個單位約等於7像素（在96 DPI下）
            total_width_px = 0
            for i in range(1, 15):
                col_letter = get_column_letter(i)
                col_width = ws.column_dimensions[col_letter].width
                if col_width is None:
                    col_width = 8.43  # Excel默認列寬
                total_width_px += int(round(col_width * 7))
            
            # 計算列高（像素）
            # Excel行高單位轉換：1個單位約等於1.33像素（在96 DPI下）
            row_height = ws.row_dimensions[row].height
            if row_height is None:
                row_height = 15  # Excel默認行高
            row_height_px = int(round(row_height * 1.33))
            
            # 計算置中偏移（水平居中）
            x_offset_px = max(0, (total_width_px - img.width) // 2)
            # 計算置中偏移（垂直居中）
            y_offset_px = max(0, (row_height_px - img.height) // 2)
            
            # 找到起始欄位（從A列開始累積寬度，找到x_offset_px落在哪一列）
            accum = 0
            start_col = 1
            col_off_px = x_offset_px
            for col_idx in range(1, 15):
                col_letter = get_column_letter(col_idx)
                w = ws.column_dimensions[col_letter].width or 8.43
                cw_px = int(round(w * 7))
                if accum + cw_px >= x_offset_px:
                    start_col = col_idx
                    col_off_px = x_offset_px - accum
                    break
                accum += cw_px
            
            # 建立置中錨點
            anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=start_col - 1,
                    colOff=pixels_to_EMU(col_off_px),
                    row=row - 1,
                    rowOff=pixels_to_EMU(y_offset_px),
                ),
                ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
            )
            img.anchor = anchor
            ws.add_image(img)
            print(f"成功插入Logo圖片（錨點置中方式）")
            return
        except Exception as e2:
            print(f"錨點置中方式失敗: {e2}，嘗試簡單方式...")
            import traceback
            traceback.print_exc()
        
        # 回退方式：簡單插入（可能無法完全置中）
        try:
            # 重新載入圖片
            img = XLImage(logo_path)
            img.width = int(598 * 0.9)
            img.height = int(61 * 0.9)
            ws.add_image(img, f"A{row}")
            print(f"成功插入Logo圖片（簡單方式，可能未完全置中）")
        except Exception as e3:
            print(f"所有圖片插入方式都失敗: {e3}")
            import traceback
            traceback.print_exc()
                
    except Exception as e:
        print(f"插入Logo圖片時發生錯誤: {e}")
        import traceback
        traceback.print_exc()

def generate_trade_notice_template(
    cus_id:str,
    cus_info:pd.DataFrame,
    df_today_trade:pd.DataFrame,
    df_today_clearing_money:pd.DataFrame,
    tday_str:str,
    df_today_bargain:pd.DataFrame,
    tday_plus_1:str,
    tday_plus_2:str,
    ifbankok:str):
    """
    產生『可轉換資產交換選擇權 交易通知』Excel 模板。
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "交易通知"
        
        # 設定頁面和欄寬
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.scale = 60
        _set_col_widths(ws)

        # === 頁首區域 ===
        # Logo
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        ws.row_dimensions[1].height = 52
        _add_logo(ws, r"\\10.72.228.112\cbas業務公用區\CBAS_Trading_Maker\figs\unipsg.png", row=1)

        # 公司名稱
        ws.row_dimensions[2].height = 24
        _write_merged_cell(ws, 2, 1, 14, "統一綜合證券股份有限公司 計量交易部", 16, True, "center")

        # 主標題
        _write_merged_cell(ws, 3, 1, 14, "可轉換資產交換選擇權 交易通知", 15, True, "center")

        # 客戶問候 + 日期
        cus_name = cus_info["CUSNAME"].iat[0] if isinstance(cus_info, pd.DataFrame) and not cus_info.empty else ""
        if len(cus_name) >= 3:
            cus_name = cus_name[0] + "Ｏ" + cus_name[-1]
        elif len(cus_name) == 2:
            cus_name = cus_name[0] + "Ｏ"

        greet = f"貴客戶 {cus_name} 您好" if cus_name else "貴客戶 您好"
        _write_merged_cell(ws, 4, 1, 12, greet, 14, False, "left")
        _write_merged_cell(ws, 4, 13, 14, f"通知日期：{datetime.now():%Y/%m/%d}", 13, False, "right")

        # === 一、新作 ===
        row = 6
        row = _section_title(ws, row, "一、資產交換選擇權－新作")
        row += 1

        df_buy = df_today_trade[df_today_trade['新作契約編號'].notna()]
        if not df_buy.empty:
            headers_buy = ["新作契約編號", "交易日", "交割日", "CB代號", "CB名稱", "百元價", 
                          "履約利率", "選擇權到期日", "成交張數", "成交均價", "單位權利金", "權利金總額"]
            df_buy = df_buy[headers_buy]
            
            # 格式化数据，对"權利金"列添加千分位，对"履約利率"添加%
            rows_data = []
            for row_values in df_buy.values.tolist():
                formatted_row = []
                for i, v in enumerate(row_values):
                    if isinstance(v, float) and np.isnan(v):
                        formatted_row.append(None)
                    elif headers_buy[i] == "權利金總額":
                        formatted_row.append(_format_number_with_comma(v))
                    elif headers_buy[i] == "履約利率":
                        formatted_row.append(f"{v}%" if v is not None and not pd.isna(v) else v)
                    else:
                        formatted_row.append(v)
                rows_data.append(formatted_row)
            
            row = _write_table(ws, row, headers_buy, rows_data, LIGHT_BLUE_FILL)
        else:
            _write_merged_cell(ws, row, 1, 14, "無", 13)
            row += 2

        # === 二、提解 ===
        row = _section_title(ws, row, "二、資產交換選擇權－提解")
        row += 1

        df_sell = df_today_trade[df_today_trade['解約契約編號'].notna()]
        df_sell['交割日_賣出'] = df_sell['交割日_賣出'].apply(lambda x: x.strftime('%Y%m%d') if not pd.isna(x) else "")
        if not df_sell.empty:
            headers_sell = ["解約契約編號", "交易日_賣出", "交割日_賣出", "CB代號", "CB名稱", "履約方式", 
                           "履約利率_賣出", "履約張數", "剩餘張數", "成交均價_賣出", "履約價", "交割金額", "履約損益", "原單契約編號"]
            df_sell = df_sell[headers_sell]
            df_sell.rename(columns={
                    '交易日_賣出': '交易日',
                    '交割日_賣出': '交割日',
                    '成交均價_賣出': '成交均價',
                    '履約利率_賣出': '履約利率',
                }, inplace=True)
            # 格式化数据，对"交割金額"列添加千分位，对"履約利率"添加%
            headers_sell = ["解約契約編號", "交易日", "交割日", "CB代號", "CB名稱", "履約方式", 
                           "履約利率", "履約張數", "剩餘張數", "成交均價", "履約價", "交割金額", "履約損益", "原單契約編號"]
            #df_sell = df_sell[headers_sell]
            rows_data = []
            for row_values in df_sell.values.tolist():
                formatted_row = []
                for i, v in enumerate(row_values):
                    if isinstance(v, float) and np.isnan(v):
                        formatted_row.append(None)
                    elif headers_sell[i] == "交割金額" or headers_sell[i] == "履約損益":
                        formatted_row.append(_format_number_with_comma(v))
                    elif headers_sell[i] == "履約利率":
                        formatted_row.append(f"{v}%" if v is not None and not pd.isna(v) else v)
                    else:
                        formatted_row.append(v)
                rows_data.append(formatted_row)
            
            row = _write_table(ws, row, headers_sell, rows_data, LIGHT_BLUE_FILL)
        else:
            _write_merged_cell(ws, row, 1, 14, "無", 13)
            row += 2

        # === 三、議價交易 ===
        row = _section_title(ws, row, "三、營業處所議價交易")
        row += 1
        df_today_bargain['買/賣'] = np.where(df_today_bargain['買/賣'] == 'B', '客戶賣出', '客戶買進')

        if not df_today_bargain.empty:
            headers_bargain = ['單據編號', '成交日', '交割日', 'CB代號', 'CB名稱', '買/賣',	'議價價格',	'議價張數',	'議價金額']
            headers_used = [h for h in headers_bargain if h in df_today_bargain.columns]
            
            # 格式化数据，对"議價價格"和"議價金額"列添加千分位
            rows_data = []
            for row_values in df_today_bargain[headers_used].values.tolist():
                formatted_row = []
                for i, v in enumerate(row_values):
                    if isinstance(v, float) and np.isnan(v):
                        formatted_row.append(None)
                    elif headers_used[i] == "議價金額":
                        formatted_row.append(_format_number_with_comma(v))
                    else:
                        formatted_row.append(v)
                rows_data.append(formatted_row)
            
            row = _write_table(ws, row, headers_used, rows_data, LIGHT_BLUE_FILL)
        else:
            _write_merged_cell(ws, row, 1, 14, "無", 13)
            row += 2
        
        # === 交割資訊 ===
        print(df_today_clearing_money)
        # 计算三个日期的交割金额
        tday_clearing_money = int(round(
            -df_today_clearing_money[(df_today_clearing_money['SETDATE'] == tday_str)]['PREMTOT'].sum() + 
            df_today_clearing_money[(df_today_clearing_money['SETDAT'] == tday_str)]['Adj_MTHAMT'].sum() +
            df_today_clearing_money[(df_today_clearing_money['DUEPAYDT'] == tday_str)]['SETTTOT'].sum()
        ))
        
        tday_plus_1_clearing_money = int(round(
            -df_today_clearing_money[(df_today_clearing_money['SETDATE'] == tday_plus_1)]['PREMTOT'].sum() + 
            df_today_clearing_money[(df_today_clearing_money['SETDAT'] == tday_plus_1)]['Adj_MTHAMT'].sum() +
            df_today_clearing_money[(df_today_clearing_money['DUEPAYDT'] == tday_plus_1)]['SETTTOT'].sum()
        ))
        
        tday_plus_2_clearing_money = int(round(
            -df_today_clearing_money[(df_today_clearing_money['SETDATE'] == tday_plus_2)]['PREMTOT'].sum() + 
            df_today_clearing_money[(df_today_clearing_money['SETDAT'] == tday_plus_2)]['Adj_MTHAMT'].sum() +
            df_today_clearing_money[(df_today_clearing_money['DUEPAYDT'] == tday_plus_2)]['SETTTOT'].sum()
        ))

        # 日期格式转换函数：20251002 -> 2025/10/02
        def format_date(date_str):
            if len(date_str) == 8:
                return f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
            return date_str

        row = _section_title(ws, row, "交割資訊如下")
        
        # 构建交割资讯行（可能有1-3行）
        settle_rows = []
        
        # T日
        if tday_clearing_money != 0:
            if tday_clearing_money > 0:
                settle_rows.append([format_date(tday_str), f"{abs(tday_clearing_money):,}", "客戶收款", ""])
            else:
                payment_method = "自動扣款" if ifbankok == 'Y' else "客戶匯款"
                settle_rows.append([format_date(tday_str), f"{abs(tday_clearing_money):,}", "客戶付款", payment_method])
        
        # T+1日
        if tday_plus_1_clearing_money != 0:
            if tday_plus_1_clearing_money > 0:
                settle_rows.append([format_date(tday_plus_1), f"{abs(tday_plus_1_clearing_money):,}", "客戶收款", ""])
            else:
                payment_method = "自動扣款" if ifbankok == 'Y' else "客戶匯款"
                settle_rows.append([format_date(tday_plus_1), f"{abs(tday_plus_1_clearing_money):,}", "客戶付款", payment_method])
        
        # T+2日
        if tday_plus_2_clearing_money != 0:
            if tday_plus_2_clearing_money > 0:
                settle_rows.append([format_date(tday_plus_2), f"{abs(tday_plus_2_clearing_money):,}", "客戶收款", ""])
            else:
                payment_method = "自動扣款" if ifbankok == 'Y' else "客戶匯款"
                settle_rows.append([format_date(tday_plus_2), f"{abs(tday_plus_2_clearing_money):,}", "客戶付款", payment_method])
        
        row = _write_table(ws, row, ["交易日期", "交割總額", "收付", "交割方式"], settle_rows, LIGHT_BLUE_FILL)
        
        # 加大"交割總額"列的寬度（B列，第2列）
        ws.column_dimensions['B'].width = 18

        # 建立交割資訊DataFrame（包含客戶ID和客戶名稱）
        df_clearing_money_consolidate = pd.DataFrame(settle_rows, columns=["交易日期", "交割總額", "收付", "交割方式"])
        # 添加客戶ID和客戶名稱欄位
        if isinstance(cus_info, pd.DataFrame) and not cus_info.empty:
            cus_name_full = cus_info.iloc[0].get('CUSNAME', '')
            if not df_clearing_money_consolidate.empty:
                df_clearing_money_consolidate.insert(0, '客戶ID', cus_id)
                df_clearing_money_consolidate.insert(1, '客戶名稱', cus_name_full)
            else:
                # 即使沒有交割資訊，也建立包含客戶ID和客戶名稱的空DataFrame
                df_clearing_money_consolidate = pd.DataFrame(columns=['客戶ID', '客戶名稱', '交易日期', '交割總額', '收付', '交割方式'])

        # === 備註說明 ===
        notes = (
            "交割方式為『自動扣款』者，請於上述交割日早上 8:30 前確保帳戶有足夠餘額；若回傳額不足或扣款失敗，客戶需在本公司通知後，"
            "自行匯款到統一證券交割銀行。\n"
            "交割方式為『匯款』者，請於交割日前一個營業日上午 10:00 前，使用客戶戶交易銀行自行匯款到統一證券交割銀行。\n\n"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=14)
        cell = ws.cell(row=row, column=1, value=notes)
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        cell.font = Font(name=KAI_FONT_NAME, size=12)
        cell.fill = WHITE_FILL
        row += 2
        
        # 空白行
        row += 2

        # 統一證券交割資訊
        psc_info = (
            "統一證券交割資訊：(建議設定約定轉帳)\n"
            "戶　　　　名： 統一綜合證券股份有限公司\n"
            "交割銀行帳號： 國泰世華商業銀行復興分行\n"
            "               82500+您的身分證字號末9碼(共14碼)\n"
            "交割集保帳號： 統一綜合證券\n"
            "               585T8888881\n\n"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 7, end_column=14)
        cell = ws.cell(row=row, column=1, value=psc_info)
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        cell.font = Font(name=KAI_FONT_NAME, size=12)
        cell.fill = WHITE_FILL
        row += 8

        # 客戶交割資訊
        cus_info_text = ""
        if isinstance(cus_info, pd.DataFrame) and not cus_info.empty:
            try:
                c = cus_info.iloc[0]
                # 遮蔽账号信息
                masked_bank_account = _mask_account_number(str(c.get('BNKACTNO', '')))
                masked_center_account = _mask_account_number(str(c.get('CENTERNO', '')))
                
                cus_info_text = (
                    f"客戶交割資訊：\n"
                    f"戶　　　　名： {cus_name}\n"
                    f"交割銀行帳號： {c.get('BNKNAME','')} {c.get('BNKBRH','')}\n"
                    f"               {masked_bank_account}\n"
                    f"交割集保帳號： {masked_center_account}\n\n"
                )
            except Exception as e:
                print(f"客戶交割資訊錯誤: {e}")
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 7, end_column=14)
        cell = ws.cell(row=row, column=1, value=cus_info_text)
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        cell.font = Font(name=KAI_FONT_NAME, size=12)
        cell.fill = WHITE_FILL
        row += 8

        # === Footer ===
        footer_row = row + 2
        footer = (
            "本交易通知如有任何問題，請聯繫本公司:\n"
            "02-27463801、02-27463670\n"
            "Email：PSC.CBAS@uni-psg.com\n"
            "Line ID：@psc.cbas\n"
            "統一綜合證券祝您投資順利！"
        )
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=14)
        cell = ws.cell(row=footer_row, column=1, value=footer)
        cell.font = Font(name=KAI_FONT_NAME, size=12)
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        cell.fill = WHITE_FILL
        ws.row_dimensions[footer_row].height = 80

        # === 全頁白色背景（保留淺藍表頭）===
        for r in range(1, ws.max_row + 1):
            for c in range(1, 15):
                cell = ws.cell(row=r, column=c)
                if cell.fill and getattr(cell.fill, 'fgColor', None):
                    if cell.fill.fgColor.rgb in (LIGHT_BLUE_FILL.fgColor.rgb, 'E6F2FF'):
                        continue
                cell.fill = WHITE_FILL

        yyyy_mm_dd = datetime.today().strftime('%Y-%m-%d')
        
        # 正確處理客戶姓名
        if isinstance(cus_info, pd.DataFrame) and not cus_info.empty:
            cusname_full = cus_info.iloc[0].get('CUSNAME', '')
            cusemail_full = cus_info.iloc[0].get('EMAIL', '')
        else:
            cusname_full = ''
            cusemail_full = ''
        
        # 姓名遮蔽處理
        if len(cusname_full) >= 3:
            cusName = cusname_full[0] + 'Ｏ' + cusname_full[-1]  # 張三丰 -> 張Ｏ丰
        elif len(cusname_full) == 2:
            cusName = cusname_full[0] + 'Ｏ'  # 張三 -> 張Ｏ
        elif len(cusname_full) == 1:
            cusName = cusname_full  # 王 -> 王
        else:
            cusName = '客戶'  # 空名稱
        
        cusEmail = cusemail_full[:4] if cusemail_full else ''
        fileName = f"{yyyy_mm_dd}_統一證券CBAS成交明細_{cusName}_{cusEmail}.xlsx"
        
        # 建立目錄路徑
        excel_dir = os.path.join(trade_notice_dir, tday_str, 'Excel')
        pdf_dir = os.path.join(trade_notice_dir, tday_str, 'PDF')
        os.makedirs(excel_dir, exist_ok=True)
        os.makedirs(pdf_dir, exist_ok=True)
        
        excel_output_path = os.path.join(excel_dir, fileName)
        wb.save(excel_output_path)
        
        # 轉換為PDF
        pdf_file_path = _convert_excel_to_pdf(excel_output_path, pdf_dir)
        
        return excel_output_path, pdf_file_path, df_clearing_money_consolidate
        
    except Exception as e:
        print(f"生成交易通知模板時發生錯誤: {e} (行: {e.__traceback__.tb_lineno if e.__traceback__ else '?'})")
        return None, None, pd.DataFrame()

if __name__ == "__main__":
    excel_path, pdf_path = generate_trade_notice_template()
    #if excel_path:
    #    print(f"Excel模板已輸出：{os.path.abspath(excel_path)}")
    if pdf_path:
        print(f"PDF模板已輸出：{os.path.abspath(pdf_path)}")
